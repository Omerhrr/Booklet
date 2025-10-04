# app/crud/reports.py
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_
from .. import models, crud
from datetime import date, timedelta
from typing import Optional, List, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from weasyprint import HTML
from fastapi.templating import Jinja2Templates
from .ledger import get_profit_and_loss_data
from fastapi.encoders import jsonable_encoder

def get_sales_report(db: Session, business_id: int, start_date: date, end_date: date, customer_id: Optional[int] = None, branch_id: Optional[int] = None):
    """
    Fetches sales invoices for a report, with optional filters.
    """
    query = db.query(models.SalesInvoice).filter(
        models.SalesInvoice.business_id == business_id,
        models.SalesInvoice.invoice_date.between(start_date, end_date)
    ).options(
        joinedload(models.SalesInvoice.customer),
        joinedload(models.SalesInvoice.branch)
    ).order_by(models.SalesInvoice.invoice_date.desc())

    if customer_id:
        query = query.filter(models.SalesInvoice.customer_id == customer_id)
    if branch_id:
        query = query.filter(models.SalesInvoice.branch_id == branch_id)
        
    invoices = query.all()
    total_sales = sum(inv.total_amount for inv in invoices)
    
    return invoices, total_sales


def get_purchase_report(db: Session, business_id: int, start_date: date, end_date: date, vendor_id: Optional[int] = None, branch_id: Optional[int] = None):
    """
    Fetches purchase bills for a report, with optional filters.
    """
    query = db.query(models.PurchaseBill).filter(
        models.PurchaseBill.business_id == business_id,
        models.PurchaseBill.bill_date.between(start_date, end_date)
    ).options(
        joinedload(models.PurchaseBill.vendor),
        joinedload(models.PurchaseBill.branch)
    ).order_by(models.PurchaseBill.bill_date.desc())

    if vendor_id:
        query = query.filter(models.PurchaseBill.vendor_id == vendor_id)
    if branch_id:
        query = query.filter(models.PurchaseBill.branch_id == branch_id)
        
    bills = query.all()
    total_purchases = sum(bill.total_amount for bill in bills)
    
    return bills, total_purchases



def get_expense_report(db: Session, business_id: int, start_date: date, end_date: date, category: Optional[str] = None, branch_id: Optional[int] = None):
    """
    Fetches expenses for a report, with optional filters.
    """
    query = db.query(models.Expense).filter(
        models.Expense.business_id == business_id,
        models.Expense.expense_date.between(start_date, end_date)
    ).options(
        joinedload(models.Expense.branch),
        joinedload(models.Expense.vendor)
    ).order_by(models.Expense.expense_date.desc())

    if category:
        query = query.filter(models.Expense.category == category)
    if branch_id:
        query = query.filter(models.Expense.branch_id == branch_id)
        
    expenses = query.all()
    total_expenses = sum(exp.amount for exp in expenses)
    
    return expenses, total_expenses


def get_trial_balance_data(db: Session, business_id: int, as_of_date: date, branch_id: Optional[int] = None):
    """
    Calculates the balance of every account for a professional, grouped Trial Balance report.
    Can be filtered by branch.
    """
    query = db.query(models.Account).filter(models.Account.business_id == business_id)
    all_accounts = query.order_by(models.Account.type, models.Account.name).all()


    report_data = {
        "Asset": {"lines": [], "total_debit": 0.0, "total_credit": 0.0},
        "Liability": {"lines": [], "total_debit": 0.0, "total_credit": 0.0},
        "Equity": {"lines": [], "total_debit": 0.0, "total_credit": 0.0},
        "Revenue": {"lines": [], "total_debit": 0.0, "total_credit": 0.0},
        "Expense": {"lines": [], "total_debit": 0.0, "total_credit": 0.0},
        "grand_total_debit": 0.0,
        "grand_total_credit": 0.0
    }

    for acc in all_accounts:
        debit_query = db.query(func.sum(models.LedgerEntry.debit)).filter(
            models.LedgerEntry.account_id == acc.id,
            models.LedgerEntry.transaction_date <= as_of_date
        )
        credit_query = db.query(func.sum(models.LedgerEntry.credit)).filter(
            models.LedgerEntry.account_id == acc.id,
            models.LedgerEntry.transaction_date <= as_of_date
        )
        
        if branch_id: 
            debit_query = debit_query.filter(models.LedgerEntry.branch_id == branch_id)
            credit_query = credit_query.filter(models.LedgerEntry.branch_id == branch_id)

        debit_sum = debit_query.scalar() or 0.0
        credit_sum = credit_query.scalar() or 0.0
        
        balance = debit_sum - credit_sum
        if balance == 0:
            continue

        debit_balance, credit_balance = 0.0, 0.0
        if acc.type in [models.AccountType.ASSET, models.AccountType.EXPENSE]:
            debit_balance = balance if balance > 0 else 0.0
            credit_balance = -balance if balance < 0 else 0.0
        else: # Liability, Equity, Revenue
            credit_balance = -balance if balance < 0 else balance
            debit_balance = balance if balance > 0 else 0.0

        line_data = {
            "account_name": acc.name,
            "debit": debit_balance,
            "credit": credit_balance
        }
        
        acc_type_str = acc.type.value
        report_data[acc_type_str]["lines"].append(line_data)
        report_data[acc_type_str]["total_debit"] += debit_balance
        report_data[acc_type_str]["total_credit"] += credit_balance
        report_data["grand_total_debit"] += debit_balance
        report_data["grand_total_credit"] += credit_balance

    return report_data

def export_to_excel(headers: List[str], data: List[List[Any]], report_title: str) -> BytesIO:
    """
    Generic function to export a list of lists into an Excel file in memory.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_title

    sheet.append([report_title])
    sheet.append(headers)
    
    title_cell = sheet['A1']
    title_cell.font = Font(bold=True, size=16)
    
    for i in range(1, len(headers) + 1):
        col_letter = get_column_letter(i)
        sheet[f'{col_letter}2'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 20

    for row_data in data:
        sheet.append(row_data)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return buffer



def get_ar_aging_report(db: Session, business_id: int, branch_id: int, as_of_date: date):
    """
    Generates an Accounts Receivable aging report, NOW FILTERED BY BRANCH.
    """
    unpaid_invoices = db.query(models.SalesInvoice).filter(
        models.SalesInvoice.business_id == business_id,
        models.SalesInvoice.branch_id == branch_id, 
        or_(
            models.SalesInvoice.status == 'Unpaid',
            models.SalesInvoice.status == 'Partially Paid'
        )
    ).options(
        joinedload(models.SalesInvoice.customer)
    ).all()

    report = {
        'current': {'invoices': [], 'total': 0.0},
        '1-30': {'invoices': [], 'total': 0.0},
        '31-60': {'invoices': [], 'total': 0.0},
        '61-90': {'invoices': [], 'total': 0.0},
        '90+': {'invoices': [], 'total': 0.0},
        'grand_total': 0.0
    }

    for inv in unpaid_invoices:
        balance_due = inv.total_amount - inv.paid_amount
        report['grand_total'] += balance_due
        
        if inv.due_date >= as_of_date:
            bucket = 'current'
        else:
            days_overdue = (as_of_date - inv.due_date).days
            if 1 <= days_overdue <= 30:
                bucket = '1-30'
            elif 31 <= days_overdue <= 60:
                bucket = '31-60'
            elif 61 <= days_overdue <= 90:
                bucket = '61-90'
            else:
                bucket = '90+'
        
        report[bucket]['invoices'].append(inv)
        report[bucket]['total'] += balance_due
        
    return report

def get_ap_aging_report(db: Session, business_id: int, branch_id: int, as_of_date: date):
    """
    Generates an Accounts Payable aging report, NOW FILTERED BY BRANCH.
    """
    unpaid_bills = db.query(models.PurchaseBill).filter(
        models.PurchaseBill.business_id == business_id,
        models.PurchaseBill.branch_id == branch_id,
        or_(
            models.PurchaseBill.status == 'Unpaid',
            models.PurchaseBill.status == 'Partially Paid'
        )
    ).options(
        joinedload(models.PurchaseBill.vendor)
    ).all()

    report = {
        'current': {'bills': [], 'total': 0.0},
        '1-30': {'bills': [], 'total': 0.0},
        '31-60': {'bills': [], 'total': 0.0},
        '61-90': {'bills': [], 'total': 0.0},
        '90+': {'bills': [], 'total': 0.0},
        'grand_total': 0.0
    }

    for bill in unpaid_bills:
        balance_due = bill.total_amount - bill.paid_amount
        report['grand_total'] += balance_due

        bill_due_date = bill.due_date

        if bill_due_date >= as_of_date:
            bucket = 'current'
        else:
            days_overdue = (as_of_date - bill_due_date).days
            if 1 <= days_overdue <= 30:
                bucket = '1-30'
            elif 31 <= days_overdue <= 60:
                bucket = '31-60'
            elif 61 <= days_overdue <= 90:
                bucket = '61-90'
            else:
                bucket = '90+'
        
        report[bucket]['bills'].append(bill)
        report[bucket]['total'] += balance_due
        
    return report


def get_dashboard_data(db: Session, business_id: int, branch_id: int):
    """
    Gathers all 16+ data points for the robust, permissioned dashboard.
    This is the fully corrected and complete version.
    """
    today = date.today()
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)

    # --- Financial Health KPIs (1-4) ---
    ar_report = crud.reports.get_ar_aging_report(db, business_id, branch_id, today)
    ap_report = crud.reports.get_ap_aging_report(db, business_id, branch_id, today)
    
    payment_accounts = crud.banking.get_payment_accounts(db, business_id, branch_id)
    cash_balance = 0
    if payment_accounts:
        cash_balance = db.query(func.sum(models.LedgerEntry.debit - models.LedgerEntry.credit)).filter(
            models.LedgerEntry.account_id.in_([acc.id for acc in payment_accounts]),
            models.LedgerEntry.branch_id == branch_id
        ).scalar() or 0.0

    pnl_ytd = crud.reports.get_profit_and_loss_data(db, business_id, start_of_year, today, branch_id)

    # --- Sales Performance KPIs (5-8) ---
    sales_mtd = db.query(func.sum(models.SalesInvoice.total_amount)).filter(
        models.SalesInvoice.branch_id == branch_id,
        models.SalesInvoice.invoice_date.between(start_of_month, today)
    ).scalar() or 0.0

    new_customers_mtd = db.query(func.count(models.Customer.id)).filter(
        models.Customer.branch_id == branch_id,
        models.Customer.created_at >= start_of_month
    ).scalar() or 0

    sales_invoices_ytd = db.query(models.SalesInvoice).filter(
        models.SalesInvoice.branch_id == branch_id,
        models.SalesInvoice.invoice_date.between(start_of_year, today)
    ).all()
    avg_invoice_value = sum(inv.total_amount for inv in sales_invoices_ytd) / len(sales_invoices_ytd) if sales_invoices_ytd else 0.0

    top_selling_product = db.query(
        models.Product.name,
        func.sum(models.SalesInvoiceItem.quantity * models.SalesInvoiceItem.price).label('total_revenue')
    ).join(models.SalesInvoiceItem, models.SalesInvoiceItem.product_id == models.Product.id)\
     .join(models.SalesInvoice, models.SalesInvoice.id == models.SalesInvoiceItem.sales_invoice_id)\
     .filter(
        models.SalesInvoice.branch_id == branch_id,
        models.SalesInvoice.invoice_date.between(start_of_month, today)
    ).group_by(models.Product.name).order_by(func.sum(models.SalesInvoiceItem.quantity * models.SalesInvoiceItem.price).desc()).first()

    # --- Purchase & Expense KPIs (9-12) ---
    purchases_mtd = db.query(func.sum(models.PurchaseBill.total_amount)).filter(
        models.PurchaseBill.branch_id == branch_id,
        models.PurchaseBill.bill_date.between(start_of_month, today)
    ).scalar() or 0.0

    expenses_mtd = db.query(func.sum(models.Expense.amount)).filter(
        models.Expense.branch_id == branch_id,
        models.Expense.expense_date.between(start_of_month, today)
    ).scalar() or 0.0

    top_expense_category = db.query(
        models.Expense.category,
        func.sum(models.Expense.amount).label('total')
    ).filter(
        models.Expense.branch_id == branch_id,
        models.Expense.expense_date.between(start_of_month, today)
    ).group_by(models.Expense.category).order_by(func.sum(models.Expense.amount).desc()).first()
    
    bills_overdue_total = (ap_report.get('1-30', {}).get('total', 0) +
                           ap_report.get('31-60', {}).get('total', 0) +
                           ap_report.get('61-90', {}).get('total', 0) +
                           ap_report.get('90+', {}).get('total', 0))

    # --- List Views (15-16) ---
    recent_transactions_obj = db.query(models.LedgerEntry).options(joinedload(models.LedgerEntry.account)).filter(
        models.LedgerEntry.branch_id == branch_id
    ).order_by(models.LedgerEntry.id.desc()).limit(5).all()
    
    recent_transactions_list = [
        {
            "description": txn.description,
            "date": txn.transaction_date.strftime('%Y-%m-%d'),
            "account_name": txn.account.name,
            "debit": txn.debit,
            "credit": txn.credit
        } for txn in recent_transactions_obj
    ]

    top_unpaid_invoices_obj = db.query(models.SalesInvoice).options(joinedload(models.SalesInvoice.customer)).filter(
        models.SalesInvoice.branch_id == branch_id,
        or_(models.SalesInvoice.status == 'Unpaid', models.SalesInvoice.status == 'Partially Paid')
    ).order_by(models.SalesInvoice.due_date.asc()).limit(5).all()

    top_unpaid_invoices_list = [
        {
            "customer_name": inv.customer.name,
            "due_date": inv.due_date.strftime('%d %b, %Y'),
            "balance": inv.total_amount - inv.paid_amount
        } for inv in top_unpaid_invoices_obj
    ]

    # --- Chart Data (13-14 + new charts) ---
    
    # THE FIX: Initialize all lists before the loops
    sales_purchases_labels, sales_data, purchases_data = [], [], []
    income_vs_expense_labels, income_data, expense_data_chart = [], [], []

    for i in range(5, -1, -1):
        month_date = today - relativedelta(months=i)
        month_label = month_date.strftime('%b')
        sales_purchases_labels.append(month_label)
        income_vs_expense_labels.append(month_label)
        
        start_of_month_chart = month_date.replace(day=1)
        end_of_month_chart = (start_of_month_chart + relativedelta(months=1)) - timedelta(days=1)
        
        # Data for Sales vs. Purchases Chart
        monthly_sales = db.query(func.sum(models.SalesInvoice.total_amount)).filter(
            models.SalesInvoice.branch_id == branch_id, 
            models.SalesInvoice.invoice_date.between(start_of_month_chart, end_of_month_chart)
        ).scalar() or 0.0
        sales_data.append(round(monthly_sales, 2))

        monthly_purchases = db.query(func.sum(models.PurchaseBill.total_amount)).filter(
            models.PurchaseBill.branch_id == branch_id, 
            models.PurchaseBill.bill_date.between(start_of_month_chart, end_of_month_chart)
        ).scalar() or 0.0
        purchases_data.append(round(monthly_purchases, 2))

        # Data for Income vs. Expense Chart
        monthly_income = db.query(func.sum(models.LedgerEntry.credit - models.LedgerEntry.debit)).join(models.Account).filter(
            models.Account.type == models.AccountType.REVENUE,
            models.LedgerEntry.branch_id == branch_id,
            models.LedgerEntry.transaction_date.between(start_of_month_chart, end_of_month_chart)
        ).scalar() or 0.0
        income_data.append(round(monthly_income, 2))

        monthly_expenses = db.query(func.sum(models.LedgerEntry.debit - models.LedgerEntry.credit)).join(models.Account).filter(
            models.Account.type == models.AccountType.EXPENSE,
            models.LedgerEntry.branch_id == branch_id,
            models.LedgerEntry.transaction_date.between(start_of_month_chart, end_of_month_chart)
        ).scalar() or 0.0
        expense_data_chart.append(round(monthly_expenses, 2))

    # Data for Expense Breakdown Pie Chart
    expense_breakdown_ytd = db.query(
        models.Expense.category, func.sum(models.Expense.amount).label('total')
    ).filter(
        models.Expense.branch_id == branch_id,
        models.Expense.expense_date.between(start_of_year, today)
    ).group_by(models.Expense.category).order_by(func.sum(models.Expense.amount).desc()).all()
    expense_pie_chart_data = [{'name': category, 'value': round(total, 2)} for category, total in expense_breakdown_ytd]

    # Data for Aging Bar Chart
    aging_chart_data = {
        "labels": ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days"],
        "receivables": [
            ar_report['current']['total'], ar_report['1-30']['total'], ar_report['31-60']['total'],
            ar_report['61-90']['total'], ar_report['90+']['total']
        ],
        "payables": [
            ap_report['current']['total'], ap_report['1-30']['total'], ap_report['31-60']['total'],
            ap_report['61-90']['total'], ap_report['90+']['total']
        ]
    }

    # --- Assemble the final payload ---
    dashboard_data = {
        "kpis": {
            "total_receivables": ar_report['grand_total'],
            "total_payables": ap_report['grand_total'],
            "cash_balance": cash_balance,
            "net_profit_ytd": pnl_ytd.get("net_profit", 0.0),
            "sales_mtd": sales_mtd,
            "new_customers_mtd": new_customers_mtd,
            "avg_invoice_value": avg_invoice_value,
            "top_selling_product": top_selling_product.name if top_selling_product else "N/A",
            "purchases_mtd": purchases_mtd,
            "expenses_mtd": expenses_mtd,
            "top_expense_category": top_expense_category.category if top_expense_category else "N/A",
            "bills_overdue": bills_overdue_total,
        },
        "lists": {
            "recent_transactions": recent_transactions_list,
            "top_unpaid_invoices": top_unpaid_invoices_list,
        },
        "charts": {
            "sales_purchases": {
                "labels": sales_purchases_labels,
                "sales": sales_data,
                "purchases": purchases_data
            },
            "expense_breakdown": expense_pie_chart_data,
            "income_vs_expense": {
                "labels": income_vs_expense_labels,
                "income": income_data,
                "expenses": expense_data_chart
            },
            "aging_summary": aging_chart_data
        }
    }
    
    return dashboard_data

def get_account_statement_data(db: Session, business_id: int, start_date: date, end_date: date, customer_id: Optional[int] = None, vendor_id: Optional[int] = None):
    """
    Fetches all transactions for a customer or vendor to generate a statement.
    """
    if not customer_id and not vendor_id:
        return [], 0.0, 0.0

    if customer_id:
        target_id = customer_id
        account_name = "Accounts Receivable"
        target_model = models.Customer
        target_relation = models.LedgerEntry.customer_id
    else:
        target_id = vendor_id
        account_name = "Accounts Payable"
        target_model = models.Vendor
        target_relation = models.LedgerEntry.vendor_id

    target = db.query(target_model).filter(target_model.id == target_id, target_model.business_id == business_id).first()
    if not target:
        return [], 0.0, 0.0, None

    opening_balance = 0.0
    opening_debits = db.query(func.sum(models.LedgerEntry.debit)).filter(
        target_relation == target_id,
        models.LedgerEntry.account.has(name=account_name),
        models.LedgerEntry.transaction_date < start_date
    ).scalar() or 0.0
    opening_credits = db.query(func.sum(models.LedgerEntry.credit)).filter(
        target_relation == target_id,
        models.LedgerEntry.account.has(name=account_name),
        models.LedgerEntry.transaction_date < start_date
    ).scalar() or 0.0

    if customer_id:
        opening_balance = opening_debits - opening_credits
    else:
        opening_balance = opening_credits - opening_debits

    entries = db.query(models.LedgerEntry).filter(
        target_relation == target_id,
        models.LedgerEntry.account.has(name=account_name),
        models.LedgerEntry.transaction_date.between(start_date, end_date)
    ).order_by(models.LedgerEntry.transaction_date.asc(), models.LedgerEntry.id.asc()).all()

    running_balance = opening_balance
    statement_lines = []
    for entry in entries:
        if customer_id:
            running_balance += entry.debit - entry.credit
        else:
            running_balance += entry.credit - entry.debit
        
        statement_lines.append({
            "entry": entry,
            "balance": running_balance
        })

    return statement_lines, opening_balance, running_balance, target


def render_html_to_pdf(template_path: str, context: dict, templates: Jinja2Templates) -> BytesIO:
    """
    Renders a Jinja2 template to HTML and then converts it to a PDF using WeasyPrint.
    """
    template = templates.get_template(template_path)
    html_content = template.render(context)
    
    pdf_buffer = BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    
    return pdf_buffer



def get_stock_valuation_report(db: Session, business_id: int, branch_id: Optional[int] = None):
    """
    Generates a stock valuation report based on purchase price.
    """
    query = db.query(models.Product).join(models.Branch).filter(
        models.Branch.business_id == business_id
    ).options(
        joinedload(models.Product.branch),
        joinedload(models.Product.category)
    ).order_by(models.Product.name)

    if branch_id:
        query = query.filter(models.Product.branch_id == branch_id)
    
    products = query.all()

    report_lines = []
    grand_total_value = 0.0

    for p in products:
        line_value = p.stock_quantity * p.purchase_price
        report_lines.append({
            "product_name": p.name,
            "sku": p.sku,
            "branch_name": p.branch.name,
            "category_name": p.category.name,
            "stock_quantity": p.stock_quantity,
            "purchase_price": p.purchase_price,
            "total_value": line_value
        })
        grand_total_value += line_value
        
    return report_lines, grand_total_value

def get_consolidated_dashboard_data(db: Session, business_id: int):
    """
    Gathers key performance indicators for each branch for a consolidated view.
    """
    branches = db.query(models.Branch).filter_by(business_id=business_id).all()
    
    today = date.today()
    start_of_month = today.replace(day=1)
    
    report_data = []

    for branch in branches:
        # Get P&L data for the current month for this specific branch
        pnl_data = get_profit_and_loss_data(
            db, 
            business_id=business_id, 
            start_date=start_of_month, 
            end_date=today, 
            branch_id=branch.id
        )
        
        # Get total purchases for the current month for this branch
        total_purchases = db.query(func.sum(models.PurchaseBill.total_amount)).filter(
            models.PurchaseBill.branch_id == branch.id,
            models.PurchaseBill.bill_date.between(start_of_month, today)
        ).scalar() or 0.0

        report_data.append({
            "branch_name": branch.name,
            "total_sales": pnl_data.get("total_revenue", 0.0),
            "gross_profit": pnl_data.get("gross_profit", 0.0),
            "total_purchases": total_purchases
        })
        
    return report_data





def get_vat_report_data(db: Session, business_id: int, branch_id: int, start_date: date, end_date: date):
    """
    Calculates the data needed for a VAT Return report for a specific branch and period.
    """
    # 1. Find the VAT accounts for the business
    output_vat_account = db.query(models.Account).filter_by(business_id=business_id, name="VAT Payable (Output VAT)").first()
    input_vat_account = db.query(models.Account).filter_by(business_id=business_id, name="VAT Receivable (Input VAT)").first()

    if not output_vat_account or not input_vat_account:
        raise ValueError("VAT accounts are not configured for this business.")

    # 2. Calculate Total Output VAT (Credits to the liability account)
    total_output_vat = db.query(func.sum(models.LedgerEntry.credit)).filter(
        models.LedgerEntry.account_id == output_vat_account.id,
        models.LedgerEntry.branch_id == branch_id,
        models.LedgerEntry.transaction_date.between(start_date, end_date)
    ).scalar() or 0.0

    # 3. Calculate Total Input VAT (Debits to the asset account)
    total_input_vat = db.query(func.sum(models.LedgerEntry.debit)).filter(
        models.LedgerEntry.account_id == input_vat_account.id,
        models.LedgerEntry.branch_id == branch_id,
        models.LedgerEntry.transaction_date.between(start_date, end_date)
    ).scalar() or 0.0

    # 4. Calculate the net amount payable or refundable
    net_vat_due = total_output_vat - total_input_vat

    # 5. Fetch the detailed transactions for the audit trail
    output_vat_transactions = db.query(models.LedgerEntry).filter(
        models.LedgerEntry.account_id == output_vat_account.id,
        models.LedgerEntry.branch_id == branch_id,
        models.LedgerEntry.transaction_date.between(start_date, end_date)
    ).order_by(models.LedgerEntry.transaction_date).all()

    input_vat_transactions = db.query(models.LedgerEntry).filter(
        models.LedgerEntry.account_id == input_vat_account.id,
        models.LedgerEntry.branch_id == branch_id,
        models.LedgerEntry.transaction_date.between(start_date, end_date)
    ).order_by(models.LedgerEntry.transaction_date).all()

    return {
        "total_output_vat": total_output_vat,
        "total_input_vat": total_input_vat,
        "net_vat_due": net_vat_due,
        "output_vat_transactions": output_vat_transactions,
        "input_vat_transactions": input_vat_transactions
    }












def get_business_data_as_json(db: Session, business_id: int, branch_id: int | None):
    """
    Fetches all relevant business data and serializes it into a JSON-ready
    dictionary, filtered by the selected branch if one is provided.
    """
    
    # Base queries
    customers_query = db.query(models.Customer).filter(models.Customer.business_id == business_id)
    vendors_query = db.query(models.Vendor).filter(models.Vendor.business_id == business_id)
    products_query = db.query(models.Product).join(models.Branch).filter(models.Branch.business_id == business_id)
    sales_query = db.query(models.SalesInvoice).filter(models.SalesInvoice.business_id == business_id)
    purchases_query = db.query(models.PurchaseBill).filter(models.PurchaseBill.business_id == business_id)
    expenses_query = db.query(models.Expense).filter(models.Expense.business_id == business_id)
    employees_query = db.query(models.Employee).filter(models.Employee.business_id == business_id)

    # Apply branch filter if a specific branch is selected
    if branch_id:
        customers_query = customers_query.filter(models.Customer.branch_id == branch_id)
        vendors_query = vendors_query.filter(models.Vendor.branch_id == branch_id)
        products_query = products_query.filter(models.Product.branch_id == branch_id)
        sales_query = sales_query.filter(models.SalesInvoice.branch_id == branch_id)
        purchases_query = purchases_query.filter(models.PurchaseBill.branch_id == branch_id)
        expenses_query = expenses_query.filter(models.Expense.branch_id == branch_id)
        employees_query = employees_query.filter(models.Employee.branch_id == branch_id)

    # Eager load relationships to create a comprehensive object graph
    sales_invoices = sales_query.options(
        joinedload(models.SalesInvoice.customer),
        joinedload(models.SalesInvoice.items).joinedload(models.SalesInvoiceItem.product)
    ).all()
    
    purchase_bills = purchases_query.options(
        joinedload(models.PurchaseBill.vendor),
        joinedload(models.PurchaseBill.items).joinedload(models.PurchaseBillItem.product)
    ).all()

    # Consolidate all data into a single dictionary
    business_data = {
        "customers": jsonable_encoder(customers_query.all()),
        "vendors": jsonable_encoder(vendors_query.all()),
        "products": jsonable_encoder(products_query.all()),
        "employees": jsonable_encoder(employees_query.all()),
        "sales_invoices": jsonable_encoder(sales_invoices),
        "purchase_bills": jsonable_encoder(purchase_bills),
        "expenses": jsonable_encoder(expenses_query.all()),
    }
    
    return business_data